#!C:\pythonCode
# -*- coding: utf-8 -*-
# @Time : 2022/8/16 11:27
# @Author : duweidong
# @File : utils_duweidong.py
# @Software: PyCharm

"""
将一些常用的操作写成函数，方便以后调用

Typical usage example:

a = [[1, 2 ,3, 4, 5], [4, 'd', 3, 5, 6]]
a = read_txt(txt='files/t.txt')
write_txt('files/t.txt', a)
"""

import os
import glob
import time
import shutil
import cv2
import openpyxl

from typing import List, Dict, Tuple


def read_txt(txt: str, separator: str = None) -> List[List]:
    """
    the separator is blank, or tabs:'\t'

    Args:
        txt: path of txt file
        separator:

    Returns:

    """

    texture = []
    with open(txt, 'r', encoding='utf-8') as f:
        for line in f.readlines():
            texture.append(line.strip().split(separator))

    return texture


def write_txt(txt: str, texture: List[List], separator: str = None) -> None:
    """
    the separator is blank, or tabs:'\t'

    Args:
        txt:
        texture:
        separator:

    Returns:

    """
    separator = '\t' if not separator else separator

    with open(txt, 'w', encoding='utf-8') as f:
        for line in texture:
            line = [str(item) for item in line]
            f.write(separator.join(line) + '\n')


def create_same_dir(src_root: str, dst_root: str) -> None:
    """
    递归创建相同的文件夹目录

    Args:
        src_root:
        dst_root:

    Returns:

    """

    if not os.path.isdir(src_root):
        return None

    os.makedirs(dst_root, exist_ok=True)

    for directory in os.listdir(src_root):
        src_directory = os.path.join(src_root, directory)
        dst_directory = os.path.join(dst_root, directory)
        create_same_dir(src_directory, dst_directory)


def extract_frame(video_path: str, save_path: str = None, exp_fps: int = 1) -> int:
    """
    返回当前视频保存多少张图片 保存的名字为图片是视频的第几帧

    Args:
        video_path:
        save_path:
        exp_fps: 自己设置的fps

    Returns:
        cnt: 保存的图片个数

    """

    video_name = os.path.basename(video_path).split('.')[0]
    save_path = video_path.split('.')[0] if not save_path else os.path.join(save_path, video_name)

    try:
        print(f"extracting {video_path}")

        video_capture = cv2.VideoCapture(video_path)
        fps = video_capture.get(cv2.CAP_PROP_FPS)  # 表示1s有多少帧

        if exp_fps > fps:
            raise Exception(f"video fps : {fps}\n"
                            f"expect fps : {exp_fps}\n"
                            f"expect exp_fps <= fps rather than exp_fps > fps")
        interval = round(fps / exp_fps)  # 由原fps变为现在的fps 需要间隔多少帧进行抽帧

        os.makedirs(save_path, exist_ok=True)

        ind = cnt = 0
        while True:
            success, frame = video_capture.read()
            if not success:
                break
            if ind % interval == 0:
                cnt += 1  # 保存的第几张图片
                save_name = os.path.join(save_path, "%06d" % cnt + '.jpg')
                cv2.imwrite(save_name, frame)
            ind += 1
        print(f"save {cnt} frames to {save_path}")
        return cnt  # 返回抽取的帧数

    except Exception as e:
        print(f"Failed\n"
              f"{e}")


def read_excel(excel_path: str, sheet_name: str = None) -> List[List]:
    """
    https://openpyxl.readthedocs.io/en/stable/index.html
    Args:
        excel_path:
        sheet_name:

    Returns:

    """

    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    worksheet = workbook['Sheet1' if not sheet_name else sheet_name]

    # texture = []
    # for i in range(1, worksheet.max_row):
    #     line = []
    #     for j in range(1, worksheet.max_column):
    #         ce = worksheet.cell(row=i, column=j)
    #         line.append(ce.value)
    #     texture.append(line)

    texture = [[worksheet.cell(row=i, column=j).value for j in range(1, worksheet.max_column)] for i in range(1, worksheet.max_row)]

    return texture
